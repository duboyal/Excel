import sys
import os
import time
import os.path
from tempfile import mkstemp
from shutil import move
from os import remove, close
import openpyxl
from openpyxl import Workbook
#from openpyxl.styles import Color, PatternFill, Font, Border
#from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils.indexed_list import IndexedList
#from openpyxl.styles import Color, PatternFill, Font
#from openpyxl.styles.borders import Border, Side
#from openpyxl.styles import Style
from argparse import ArgumentParser
from argparse import RawTextHelpFormatter
import numpy as np
#import csv

#*****************************#

pdbidList = [] #make this a list of directories
sumchrgList = []
pKoutList = []

#******************************#

f = open('PROT_SNASE/reduced_KevinAliSnase.tsv', 'r')
pdb = f.readlines()
           #line count check
#print 'line count for'+pdb_dir+'/ (pdb stuff)'
for line in pdb:
   # i = i+1
    col = line.split()
    pdbidList.append(col[0])
              
print pdbidList
print len(pdbidList)



#***********************************#

def checkPk(pdb_dir): #pdbdir is gonna be fed from pbdidList which will be the 
    

    if os.path.isdir('PROT_SNASE/QUICK/mcee_results_'+pdb_dir+'_nw'):# path only for if already inside the python scripts directory
 

        f = open('PROT_SNASE/QUICK/mcee_results_'+pdb_dir+'_nw/pK.out', 'r') #not just pKout do sumcharge

        print '\n'
        print '\n'
        print '\n'
        print '\n'
        print '\n'
        print '*******   HELLO   ********  WE ARE NOW LOOKING AT '+pdb_dir+'  *********************'
        print '\n'

        pK_out = f.readlines()
           #line count check
        print 'first column of '+pdb_dir+'/sum_crg.out\n'
        for line in pK_out[1:]: #from index one and on
            #i = i+1
            col = line.split()
            pKoutList.append(col[1])
            #for everyitem in col[1:]:
                #print everyitem
            if '<' in col[1] or '>' in col[1]:
                print 'dont use this data point pka out of range'
                print col[0]+' '+col[1]+'\n'
	        continue #goes to outter loop
            if 'ARG' in col[0] and float(col[1]) < 14.0:
                print 'warning pka of ' + col[0]
                print line, pdb_dir+'\n'
            if 'LYS' in col[0] and float(col[1]) < 10.0:
                #print '*warning*   in: '+pdb_dir+'\n'
                print 'in: ' +pdb_dir+'pka of '+ col[0]+' is less than 10, it is: '+col[1]+'\n'
                

            if float(col[2]) < 0.5:
                print col[0]+' '+'warning slope is less than 0.5: '+col[2]+'\n'
            if float(col[3]) > 4.0:
                print col[0]+' chi2 is higher than 4: '+col[3]+'\n'







        #print pKoutList
        print len(pKoutList)

#**********************************#

def fillsheet(wb):
	ws3 = wb.create_sheet(index=2, title='hello') 
	cell_header1 = 'B1' 
	ws3[cell_header1] = 'inhibitor'	
	cell_header1 = 'C1'
	ws3[cell_header1] = 'pdb'
	cell_header1 = 'D1'
	ws3[cell_header1] = 'protein charge with inhibitor'
	cell_header1 = 'E1'
	ws3[cell_header1] = 'protein charge apo'
	cell_header1 = 'F1'
	ws3[cell_header1] = 'inhibitor charge with prot'
	cell_header1 = 'G1'
	ws3[cell_header1] = 'inhibitor charge in soln'
	cell_header1 = 'H1'
	ws3[cell_header1] = 'most occupied conformer of inhibitor in solution'
	cell_header1 = 'I1'
	ws3[cell_header1] = 'most occupied conformer of inhibitor in protein'
	
	
	
	cell_header1 = 'J1'
	ws3[cell_header1] = 'delta charge prot'
	cell_header1 = 'K1'
	ws3[cell_header1] = 'delta charge inhibitor'
	cell_header1 = 'L1'
	ws3[cell_header1] = 'delta conformation inhibitor'
	cell_header1 = 'M1'
	ws3[cell_header1] = 'record'
	cell_header1 = 'N1'
	ws3[cell_header1] = 'Comments'

#***********************************#

def seeting_up_excel():
	wb = openpyxl.Workbook()
	
	i = 1
	sheetName = 'wvuheuhvuwhv'
	
	ws3 = wb.create_sheet(index=i, title=sheetName) 
			
	# Sheet header
	#location_of_prot_dir = sheetName + 'salah'
	cell_header1 = 'A1' 
	ws3[cell_header1] = 'kdjhksj' #location_of_prot_dir
	i += 1	

	print " bakhskajhskjh"

	fillsheet(wb)

	#this is where we call the different functions!!


	wb.save('data.xlsx')

seeting_up_excel()
